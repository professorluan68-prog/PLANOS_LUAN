import openpyxl
import os

path_bio = r"D:\planilhas\aprofundamentoembiologia.xlsx"
path_geo = r"D:\planilhas\aprofundamentoemgeografia.xlsx"

def inspect_xlsx(path, name):
    print(f"\n=================== {name} ===================")
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return
    wb = openpyxl.load_workbook(path, read_only=True)
    print("Sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames[:3]:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        # Get first few rows
        rows = list(sheet.iter_rows(values_only=True))
        print(f"Total rows: {len(rows)}")
        if len(rows) > 0:
            print("Headers:", rows[0])
            for r_idx, row in enumerate(rows[1:6], start=1):
                print(f"Row {r_idx}: {row}")

inspect_xlsx(path_bio, "BIOLOGIA")
inspect_xlsx(path_geo, "GEOGRAFIA")
